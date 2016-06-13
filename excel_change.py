# -*- coding: gbk -*-
"""
Created on Fri Apr 29 21:32:27 2016

@author: Bin
"""
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, colors


def get_data_dict():
    wb = load_workbook(filename='data.xlsx')
    ws = wb["Sheet1"]
    data_dict = {}
    for cx in range(ws.max_column):
        col_list = []
        col_name = ws.cell(row=1, column=cx + 1).value
        for rx in range(2, ws.max_row):
            col_list.append(ws.cell(row=rx, column=cx + 1).value)
        data_dict[col_name] = col_list
    return data_dict


def target_data_match(target):
    data_dict = get_data_dict()
    set_data_match = Font(color=colors.GREEN)
    set_data_dismatch = Font(color=colors.RED)
    wb = load_workbook(filename=target)
    ws = wb["Sheet1"]
    for cx in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=cx).value

        if col_name in data_dict.keys():
            for rx in range(2, ws.max_row + 1):
                if ws.cell(row=rx, column=cx).value in data_dict[col_name]:
                    ws.cell(row=rx, column=cx).font = set_data_match
                else:
                    ws.cell(row=rx, column=cx).font = set_data_dismatch
        wb.save(target)


def main():
    for path, dirs, fileList in os.walk(os.getcwd()):
        for fileName in fileList:
            if re.match('.*xlsx$', fileName, re.I) and fileName != 'data.xlsx':
                target_data_match(fileName)
if __name__ == '__main__':
    main()
