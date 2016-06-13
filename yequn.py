# -*- coding: gbk -*-
"""
Created on Fri Apr 29 21:32:27 2016

@author: Bin
"""
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, colors


def get_data():
    wb = load_workbook(filename='excel.xlsx')
    ws = wb["Sheet1"]
    data = []
    for rx in range(2, ws.max_row):
        tuhao = ws.cell(row=rx, column=2).value
        hanghao = rx
        jiage = ws.cell(row=rx, column=6).value
        data.append((tuhao, jiage, hanghao))
    return data


def process_data():
    data = get_data()
    part_dict = {}
    for a in data:
        if a[0] not in part_dict.keys() and a[0] != '':
            part_dict[a[0]] = (a[1], a[2])
        elif part_dict[a[0]][0] > a[1]:
            part_dict[a[0]] = (a[1], a[2])
    return part_dict


def write_data():
    part_dict = process_data()
    wb = load_workbook(filename='excel.xlsx')
    ws = wb["Sheet2"]
    r = 1
    for key in part_dict.keys():
        ws.cell(row=r, column=1).value = key
        ws.cell(row=r, column=2).value = part_dict[key][0]
        ws.cell(row=r, column=3).value = part_dict[key][1]
        r += 1
    wb.save('excel.xlsx')

if __name__ == '__main__':
    write_data()
